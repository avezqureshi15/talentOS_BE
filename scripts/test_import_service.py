import io
import sys
from pathlib import Path
from unittest.mock import patch
from uuid import uuid4

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook  # noqa: E402

from app.modules.hiring_requests.excel.import_service import (  # noqa: E402
    CandidateImportService,
    COLUMN_HEADERS,
    IMPORT_COLUMNS,
    _normalize_header,
)

inserted = []
published = []


def fake_create_queued_candidate(db, application_id, job_id, **kwargs):
    inserted.append((application_id, job_id, kwargs))
    return None


def fake_publish(topic, key, value):
    published.append((topic, key, value))


def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.append([COLUMN_HEADERS[k] for k in IMPORT_COLUMNS])
    for row in rows:
        ws.append([row.get(k, "") for k in IMPORT_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def run(rows):
    inserted.clear()
    published.clear()
    fake_db = object()
    with patch(
        "app.modules.hiring_requests.excel.import_service.create_queued_candidate",
        side_effect=fake_create_queued_candidate,
    ), patch(
        "app.modules.hiring_requests.excel.import_service.publish",
        side_effect=fake_publish,
    ), patch.object(
        CandidateImportService,
        "_resolve_external_job_id",
        return_value=str(uuid4()),
    ):
        return CandidateImportService(fake_db).import_candidates(uuid4(), build_workbook(rows))


good = {
    "name": "Alice",
    "email": "alice@example.com",
    "phone": "9876543210",
    "resume_url": "https://s3.example.com/alice.pdf",
    "candidate_type": "REFERRAL",
    "willing_to_relocate": "Yes",
}

assert _normalize_header(" Name * ") == "name"
assert _normalize_header("Willing to Relocate (Yes/No)") == "willing_to_relocate"
assert _normalize_header("Resume URL *") == "resume_url"

# 1. happy path + type + relocate
s = run([good])
assert s["total"] == 1 and s["imported"] == 1 and s["skipped_duplicates"] == 0 and s["failed"] == [], s
assert inserted[0][2]["candidate_type"] == "REFERRAL", inserted
assert inserted[0][2]["willing_to_relocate"] is True
assert len(published) == 1

# 2. validation errors: missing name, bad email, missing phone, missing resume
s = run([
    {"name": "", "email": "b@example.com", "phone": "1", "resume_url": "https://x/y.pdf"},
    {"name": "Bob", "email": "not-an-email", "phone": "1", "resume_url": "https://x/y.pdf"},
    {"name": "Carol", "email": "c@example.com", "phone": "", "resume_url": "https://x/y.pdf"},
    {"name": "Dave", "email": "d@example.com", "phone": "1", "resume_url": ""},
])
assert s["total"] == 4 and s["imported"] == 0 and len(s["failed"]) == 4, s
assert any("Name is required" in f["error"] for f in s["failed"])
assert any("Email is invalid" in f["error"] for f in s["failed"])
assert any("Phone is required" in f["error"] for f in s["failed"])
assert any("Resume URL is required" in f["error"] for f in s["failed"])

# 3. within-file duplicate emails skipped
s = run([good, {**good, "name": "Alice 2"}])
assert s["total"] == 2 and s["imported"] == 1 and s["skipped_duplicates"] == 1, s

# 4. empty rows skipped entirely
s = run([good, {}, {"name": "  "}])
assert s["total"] == 1 and s["imported"] == 1, s

# 5. numeric cells coerced to strings
s = run([{**good, "phone": 9822112233, "years_of_experience": 5.0}])
assert inserted[0][2]["candidate_phone"] == "9822112233", inserted
assert inserted[0][2]["years_of_experience"] == "5", inserted

# 6. default candidate_type REGULAR
s = run([{k: v for k, v in good.items() if k != "candidate_type"}])
assert inserted[0][2]["candidate_type"] == "REGULAR"

print("ALL ASSERTIONS PASSED")

# 7. template builder
from app.modules.hiring_requests.hiring_request_service import HiringRequestService  # noqa: E402

with patch.object(
    HiringRequestService,
    "get_hiring_request_by_id",
    return_value={"data": {"title": "Test Job", "external_job_id": "x"}},
):
    tpl_buf, tpl_name = CandidateImportService(object()).build_import_template(uuid4())
assert tpl_name.endswith(".xlsx")
from openpyxl import load_workbook  # noqa: E402

tpl_ws = load_workbook(tpl_buf).active
assert [c.value for c in tpl_ws[1]] == [COLUMN_HEADERS[k] for k in IMPORT_COLUMNS]
assert tpl_ws.max_row == 1
print("TEMPLATE OK")
