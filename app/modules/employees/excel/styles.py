"""Small openpyxl styling helpers for the employees import template.

Kept module-local to avoid triggering ``hiring_requests.excel.__init__``,
which eager-loads services that pull in the whole applications/forms
graph.
"""
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center")


def style_header_row(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def auto_width(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        header_len = len(str(ws.cell(row=1, column=col).value or ""))
        ws.column_dimensions[letter].width = max(18, header_len + 4)
