import io
import re
from datetime import date, datetime
from typing import Any, BinaryIO

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.logger import get_logger
from app.modules.employees.employee_model import Employee
from app.modules.employees.excel.styles import auto_width, style_header_row
from app.modules.users.user_model import User

logger = get_logger(__name__)


REQUIRED_COLUMNS = ("emp_id", "email", "name")

IMPORT_COLUMNS = (
    "emp_id",
    "email",
    "name",
    "personal_email",
    "contact_number",
    "hrms_id",
    "designation",
    "department",
    "band",
    "work_mode",
    "delivery_status",
    "work_location_type",
    "doj",
    "doe",
    "date_of_birth",
    "internship_duration",
    "skills",
    "status",
    "user_type",
)

COLUMN_HEADERS = {
    "emp_id": "Employee ID *",
    "email": "Email *",
    "name": "Name *",
    "personal_email": "Personal Email",
    "contact_number": "Contact Number",
    "hrms_id": "HRMS ID",
    "designation": "Designation",
    "department": "Department",
    "band": "Band",
    "work_mode": "Work Mode",
    "delivery_status": "Delivery Status",
    "work_location_type": "Work Location Type",
    "doj": "Date of Joining (YYYY-MM-DD)",
    "doe": "Date of Exit (YYYY-MM-DD)",
    "date_of_birth": "Date of Birth (YYYY-MM-DD)",
    "internship_duration": "Internship Duration (months)",
    "skills": "Skills",
    "status": "Status",
    "user_type": "User Type",
}

MAX_LENGTHS = {
    "emp_id": 50,
    "email": 255,
    "name": 255,
    "personal_email": 255,
    "contact_number": 20,
    "hrms_id": 100,
    "designation": 255,
    "department": 255,
    "band": 50,
    "work_mode": 50,
    "delivery_status": 50,
    "work_location_type": 50,
    "status": 50,
    "user_type": 50,
}

DATE_FIELDS = ("doj", "doe", "date_of_birth")
INT_FIELDS = ("internship_duration",)

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Header text → import key. Any raw header not in this map falls through to
# ``_normalize_header`` (lowercase, strip stars/parens, spaces→underscores).
_HEADER_ALIASES = {
    "employee_id": "emp_id",
    "employee_i_d": "emp_id",
    "hrms_i_d": "hrms_id",
    "linkedin_u_r_l": "linkedin_url",
}


def _normalize_header(raw: Any) -> str:
    value = str(raw or "").strip().lower().rstrip("*").strip()
    value = re.sub(r"\s*\(.*?\)\s*", "", value)
    key = value.replace(" ", "_")
    return _HEADER_ALIASES.get(key, key)


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _parse_date(value: str) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


class EmployeeImportService:
    """Parse an uploaded employee workbook and upsert Employee rows.

    - Dedup: within-file (by lowercased email OR emp_id) counts as a
      skipped duplicate; DB collisions (existing emp_id or email) also
      count as skipped duplicates (no overwrite).
    - Auto-link: after insert, if an existing user's email matches the new
      employee, ``users.employee_id`` is set to link the two.
    """

    def __init__(self, db: Session):
        self.db = db

    def import_employees(self, file: BinaryIO, tenant_id: int | None) -> dict:
        wb = load_workbook(file, data_only=True, read_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(values_only=True), None)
        if not header_row:
            raise HTTPException(status_code=400, detail="Workbook is empty — missing header row")

        column_indexes: dict[str, int] = {}
        for idx, raw_header in enumerate(header_row):
            key = _normalize_header(raw_header)
            if key in IMPORT_COLUMNS:
                column_indexes[key] = idx

        missing = [COLUMN_HEADERS[k] for k in REQUIRED_COLUMNS if k not in column_indexes]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required column(s) in the workbook: {', '.join(missing)}",
            )

        total = 0
        imported = 0
        skipped_duplicates = 0
        failed: list[dict] = []
        seen_emails: set[str] = set()
        seen_emp_ids: set[str] = set()

        for row_idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw = {
                key: _cell_to_str(values[column_indexes[key]]) if column_indexes[key] < len(values) else ""
                for key in IMPORT_COLUMNS
                if key in column_indexes
            }
            if not any(raw.values()):
                continue

            total += 1
            errors = self._validate_row(raw)
            if errors:
                failed.append({"row": row_idx, "error": "; ".join(errors)})
                continue

            email_key = raw["email"].strip().lower()
            emp_id_key = raw["emp_id"].strip()
            if email_key in seen_emails or emp_id_key in seen_emp_ids:
                skipped_duplicates += 1
                continue

            if self._existing_employee(email_key, emp_id_key):
                skipped_duplicates += 1
                seen_emails.add(email_key)
                seen_emp_ids.add(emp_id_key)
                continue

            try:
                self._persist(raw, tenant_id)
                self.db.commit()
            except Exception as exc:  # noqa: BLE001
                self.db.rollback()
                logger.error("Employee import row %s failed to persist: %s", row_idx, exc)
                failed.append({"row": row_idx, "error": f"Failed to save employee: {exc}"})
                continue

            seen_emails.add(email_key)
            seen_emp_ids.add(emp_id_key)
            imported += 1

        logger.info(
            "Employee import complete: tenant=%s | total=%s | imported=%s | duplicates=%s | failed=%s",
            tenant_id, total, imported, skipped_duplicates, len(failed),
        )
        return {
            "total": total,
            "imported": imported,
            "skipped_duplicates": skipped_duplicates,
            "failed": failed,
        }

    def _existing_employee(self, email_key: str, emp_id_key: str) -> Employee | None:
        return (
            self.db.query(Employee)
            .filter((Employee.email == email_key) | (Employee.emp_id == emp_id_key))
            .first()
        )

    def _persist(self, row: dict, tenant_id: int | None) -> Employee:
        employee = Employee(
            tenant_id=tenant_id,
            emp_id=row["emp_id"],
            email=row["email"].strip().lower(),
            name=row["name"],
            personal_email=row.get("personal_email") or None,
            contact_number=row.get("contact_number") or None,
            hrms_id=row.get("hrms_id") or None,
            designation=row.get("designation") or None,
            department=row.get("department") or None,
            band=row.get("band") or None,
            work_mode=row.get("work_mode") or None,
            delivery_status=row.get("delivery_status") or None,
            work_location_type=row.get("work_location_type") or None,
            doj=_parse_date(row.get("doj", "")),
            doe=_parse_date(row.get("doe", "")),
            date_of_birth=_parse_date(row.get("date_of_birth", "")),
            internship_duration=int(row["internship_duration"]) if row.get("internship_duration") else None,
            skills=row.get("skills") or None,
            status=row.get("status") or "active",
            user_type=row.get("user_type") or "employee",
        )
        self.db.add(employee)
        self.db.flush()
        # Auto-link any existing user with the same email that has no
        # employee_id yet — keeps the invariant established at invite/signup.
        self.db.query(User).filter(
            User.email == employee.email,
            User.employee_id.is_(None),
        ).update({"employee_id": employee.id}, synchronize_session=False)
        return employee

    @staticmethod
    def _validate_row(row: dict) -> list[str]:
        errors: list[str] = []

        if not row.get("emp_id"):
            errors.append("Employee ID is required")
        elif len(row["emp_id"]) > MAX_LENGTHS["emp_id"]:
            row["emp_id"] = row["emp_id"][: MAX_LENGTHS["emp_id"]]

        email = row.get("email")
        if not email:
            errors.append("Email is required")
        elif not EMAIL_RE.match(email):
            errors.append("Email is invalid")
        else:
            row["email"] = email[: MAX_LENGTHS["email"]]

        if not row.get("name"):
            errors.append("Name is required")
        elif len(row["name"]) > MAX_LENGTHS["name"]:
            row["name"] = row["name"][: MAX_LENGTHS["name"]]

        for key, limit in MAX_LENGTHS.items():
            if key in REQUIRED_COLUMNS:
                continue
            value = row.get(key)
            if value and len(value) > limit:
                row[key] = value[:limit]

        for key in DATE_FIELDS:
            value = row.get(key)
            if value and _parse_date(value) is None:
                errors.append(f"{COLUMN_HEADERS[key]} must be YYYY-MM-DD")

        for key in INT_FIELDS:
            value = row.get(key)
            if value:
                try:
                    int(value)
                except ValueError:
                    errors.append(f"{COLUMN_HEADERS[key]} must be a whole number")

        return errors

    @staticmethod
    def build_import_template() -> tuple[io.BytesIO, str]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
        for col, key in enumerate(IMPORT_COLUMNS, start=1):
            ws.cell(row=1, column=col, value=COLUMN_HEADERS[key])
        style_header_row(ws, len(IMPORT_COLUMNS))
        auto_width(ws, len(IMPORT_COLUMNS))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf, "employees_template.xlsx"
