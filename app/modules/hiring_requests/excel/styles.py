from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_body_cells(ws, row_count: int, col_count: int) -> None:
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP


def auto_width(ws, col_count: int, max_width: int = 60) -> None:
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        lengths = []
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            val = str(row[0] or "")
            lengths.append(min(len(val), max_width))
        ws.column_dimensions[letter].width = max(lengths) + 3 if lengths else 12
