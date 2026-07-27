from openpyxl import Workbook

from app.modules.hiring_requests.excel.stage_config import (
    CANDIDATE_COLUMNS,
    HIRING_REQUEST_SHEET_TITLE,
)
from app.modules.hiring_requests.excel.styles import auto_width, style_body_cells, style_header_row


def build_hiring_request_sheet(wb: Workbook, job: dict) -> None:
    ws = wb.active
    ws.title = HIRING_REQUEST_SHEET_TITLE

    fields = [
        ("Title", job.get("title", "")),
        ("Department", job.get("department", "")),
        ("Location", job.get("location", "")),
        ("Type", job.get("type", "")),
        ("Description", job.get("description", "")),
        ("Requirements", ", ".join(job.get("requirements", []) or [])),
        ("Benefits", ", ".join(job.get("benefits", []) or [])),
        ("Status", "Active" if job.get("is_active") else "Inactive"),
        ("Created At", str(job.get("created_at", "") or "")),
        ("Updated At", str(job.get("updated_at", "") or "")),
    ]

    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Value")
    style_header_row(ws, 2)

    for i, (field, value) in enumerate(fields, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)

    style_body_cells(ws, len(fields) + 1, 2)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80


def build_candidates_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict],
    columns: list[tuple[str, str]] | None = None,
) -> None:
    cols = columns or CANDIDATE_COLUMNS
    ws = wb.create_sheet(title=title)

    headers = [label for _, label in cols]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    style_header_row(ws, len(headers))

    for i, row in enumerate(rows, start=2):
        for col, (key, _) in enumerate(cols, start=1):
            value = row.get(key, "")
            ws.cell(row=i, column=col, value="" if value is None else value)

    style_body_cells(ws, len(rows) + 1, len(headers))
    auto_width(ws, len(headers))
