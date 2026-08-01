import io
import re
import uuid
from typing import Any, BinaryIO
from uuid import UUID

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.kafka import publish
from app.core.logger import get_logger
from app.modules.applications.application_repository_mutations import create_queued_candidate
from app.modules.evaluations.evaluation_schema import AsyncEvaluationMessage
from app.modules.hiring_requests.excel.styles import auto_width, style_header_row
from app.modules.hiring_requests.hiring_request_service import HiringRequestService

logger = get_logger(__name__)

REQUIRED_COLUMNS = ("name", "email", "phone", "resume_url")

IMPORT_COLUMNS = (
    "name",
    "email",
    "phone",
    "cover_letter",
    "resume_url",
    "current_ctc",
    "expected_ctc",
    "years_of_experience",
    "location",
    "notice_period",
    "how_did_you_hear",
    "linkedin_url",
    "willing_to_relocate",
    "candidate_type",
)

COLUMN_HEADERS = {
    "name": "Name *",
    "email": "Email *",
    "phone": "Phone *",
    "cover_letter": "Cover Letter",
    "resume_url": "Resume URL *",
    "current_ctc": "Current CTC",
    "expected_ctc": "Expected CTC",
    "years_of_experience": "Years of Experience",
    "location": "Location",
    "notice_period": "Notice Period",
    "how_did_you_hear": "How Did You Hear",
    "linkedin_url": "LinkedIn URL",
    "willing_to_relocate": "Willing to Relocate (Yes/No)",
    "candidate_type": "Candidate Type",
}

MAX_LENGTHS = {
    "name": 255,
    "email": 255,
    "phone": 30,
    "cover_letter": None,
    "resume_url": 1024,
    "current_ctc": 50,
    "expected_ctc": 50,
    "years_of_experience": 10,
    "location": 255,
    "notice_period": 50,
    "how_did_you_hear": 100,
    "linkedin_url": 1024,
    "candidate_type": 20,
}

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
_TRUE_VALUES = {"yes", "true", "1", "y", "yeah"}


def _normalize_header(raw: Any) -> str:
    value = str(raw or "").strip().lower().rstrip("*").strip()
    value = re.sub(r"\s*\(.*?\)\s*", "", value)
    return value.replace(" ", "_")


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _parse_relocate(value: str) -> bool:
    return value.strip().lower() in _TRUE_VALUES


class CandidateImportService:
    """Parse an uploaded candidate workbook, persist QUEUED candidates, and
    enqueue them for async AI evaluation via Kafka."""

    def __init__(self, db: Session):
        self.db = db
        self.hiring_requests = HiringRequestService(db)

    def _resolve_external_job_id(self, hiring_request_id: UUID) -> str:
        job_data = self.hiring_requests.get_hiring_request_by_id(hiring_request_id)
        external_job_id = job_data["data"].get("external_job_id")
        if not external_job_id:
            raise HTTPException(
                status_code=400,
                detail="Hiring request has no external job id — cannot import candidates",
            )
        return str(external_job_id)

    def import_candidates(self, hiring_request_id: UUID, file: BinaryIO) -> dict:
        external_job_id = self._resolve_external_job_id(hiring_request_id)

        wb = load_workbook(file, data_only=True, read_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(values_only=True), None)
        if not header_row:
            raise HTTPException(status_code=400, detail="Workbook is empty — missing header row")

        column_indexes: dict[str, int] = {}
        for idx, raw_header in enumerate(header_row):
            key = _normalize_header(raw_header)
            if key in IMPORT_COLUMNS:
                column_indexes[key] = idx

        missing = [COLUMN_HEADERS[k] for k in REQUIRED_COLUMNS if k not in column_indexes]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required column(s) in the workbook: {', '.join(missing)}",
            )

        total = 0
        imported = 0
        skipped_duplicates = 0
        failed: list[dict] = []
        seen_emails: set[str] = set()

        for row_idx, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw = {
                key: _cell_to_str(values[column_indexes[key]]) if column_indexes[key] < len(values) else ""
                for key in IMPORT_COLUMNS
                if key in column_indexes
            }
            if not any(raw.values()):
                continue

            total += 1
            row = raw
            errors = self._validate_row(row)
            if errors:
                failed.append({"row": row_idx, "error": "; ".join(errors)})
                continue

            email_key = row["email"].strip().lower()
            if email_key in seen_emails:
                skipped_duplicates += 1
                continue
            seen_emails.add(email_key)

            application_id = str(uuid.uuid4())
            try:
                create_queued_candidate(
                    self.db,
                    application_id=application_id,
                    job_id=external_job_id,
                    candidate_name=row.get("name"),
                    candidate_email=row.get("email"),
                    candidate_phone=row.get("phone"),
                    cover_letter=row.get("cover_letter") or None,
                    resume_url=row.get("resume_url"),
                    current_ctc=row.get("current_ctc") or None,
                    expected_ctc=row.get("expected_ctc") or None,
                    location=row.get("location") or None,
                    years_of_experience=row.get("years_of_experience") or None,
                    notice_period=row.get("notice_period") or None,
                    how_did_you_hear=row.get("how_did_you_hear") or None,
                    linkedin_url=row.get("linkedin_url") or None,
                    willing_to_relocate=_parse_relocate(row.get("willing_to_relocate", "")),
                    candidate_type=row.get("candidate_type") or "REGULAR",
                )
            except Exception as exc:  # noqa: BLE001
                logger.error("Import row %s failed to persist: %s", row_idx, exc)
                failed.append({"row": row_idx, "error": f"Failed to save candidate: {exc}"})
                continue

            try:
                publish(
                    topic=settings.KAFKA_TOPIC_EVALUATION_ASYNC,
                    key=application_id,
                    value=AsyncEvaluationMessage(
                        application_id=application_id,
                        job_id=external_job_id,
                        candidate_name=row.get("name"),
                        candidate_email=row.get("email"),
                        candidate_phone=row.get("phone"),
                        cover_letter=row.get("cover_letter") or None,
                        resume_url=row.get("resume_url"),
                        current_ctc=row.get("current_ctc") or None,
                        expected_ctc=row.get("expected_ctc") or None,
                        location=row.get("location") or None,
                        years_of_experience=row.get("years_of_experience") or None,
                        notice_period=row.get("notice_period") or None,
                        how_did_you_hear=row.get("how_did_you_hear") or None,
                        linkedin_url=row.get("linkedin_url") or None,
                        willing_to_relocate=_parse_relocate(row.get("willing_to_relocate", "")),
                        candidate_type=row.get("candidate_type") or "REGULAR",
                    ).model_dump_json(),
                )
            except Exception as exc:  # noqa: BLE001
                logger.error("Import row %s queued but Kafka publish failed: %s", row_idx, exc)
                failed.append({"row": row_idx, "error": "Candidate saved but failed to queue for evaluation"})
                continue

            imported += 1

        logger.info(
            "Candidate import complete: hiring_request=%s | total=%s | imported=%s | duplicates=%s | failed=%s",
            hiring_request_id, total, imported, skipped_duplicates, len(failed),
        )
        return {
            "total": total,
            "imported": imported,
            "skipped_duplicates": skipped_duplicates,
            "failed": failed,
        }

    @staticmethod
    def _validate_row(row: dict) -> list[str]:
        errors = []

        if not row.get("name"):
            errors.append("Name is required")
        else:
            row["name"] = row["name"][: MAX_LENGTHS["name"]]

        email = row.get("email")
        if not email:
            errors.append("Email is required")
        elif not EMAIL_RE.match(email):
            errors.append("Email is invalid")
        else:
            row["email"] = email[: MAX_LENGTHS["email"]]

        if not row.get("phone"):
            errors.append("Phone is required")
        else:
            row["phone"] = row["phone"][: MAX_LENGTHS["phone"]]

        if not row.get("resume_url"):
            errors.append("Resume URL is required")
        else:
            row["resume_url"] = row["resume_url"][: MAX_LENGTHS["resume_url"]]

        for key, limit in MAX_LENGTHS.items():
            if limit is None or key in ("name", "email", "phone", "resume_url"):
                continue
            value = row.get(key)
            if value and len(value) > limit:
                row[key] = value[:limit]

        if row.get("candidate_type") and len(row["candidate_type"]) > MAX_LENGTHS["candidate_type"]:
            row["candidate_type"] = row["candidate_type"][: MAX_LENGTHS["candidate_type"]]

        return errors

    def build_import_template(self, hiring_request_id: UUID) -> tuple[io.BytesIO, str]:
        job_data = self.hiring_requests.get_hiring_request_by_id(hiring_request_id)
        title = str(job_data["data"].get("title") or "import").replace("/", "_").replace("\\", "_")

        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        for col, key in enumerate(IMPORT_COLUMNS, start=1):
            ws.cell(row=1, column=col, value=COLUMN_HEADERS[key])
        style_header_row(ws, len(IMPORT_COLUMNS))
        auto_width(ws, len(IMPORT_COLUMNS))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf, f"{title}_candidates_template.xlsx"
