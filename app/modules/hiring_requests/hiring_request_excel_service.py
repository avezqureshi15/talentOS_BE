import io
from uuid import UUID

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.logger import get_logger
from app.modules.applications.application_schema import ApplicationResponse

logger = get_logger(__name__)

_APPLICATIONS_ENDPOINT: str = f"{settings.SUPABASE_FUNCTIONS_BASE_URL}/get-applications"
_TIMEOUT: int = 30

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_BODY_FONT = Font(name="Calibri", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header_row(ws, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def _style_body_cells(ws, row_count: int, col_count: int) -> None:
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _WRAP


def _auto_width(ws, col_count: int, max_width: int = 60) -> None:
    for col in range(1, col_count + 1):
        letter = get_column_letter(col)
        lengths = []
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
            val = str(row[0] or "")
            lengths.append(min(len(val), max_width))
        ws.column_dimensions[letter].width = max(lengths) + 3 if lengths else 12


def build_job_sheet(wb: Workbook, job: dict) -> None:
    ws = wb.active
    ws.title = "Job Details"

    fields = [
        ("Title", job.get("title", "")),
        ("Department", job.get("department", "")),
        ("Location", job.get("location", "")),
        ("Type", job.get("type", "")),
        ("Description", job.get("description", "")),
        ("Requirements", ", ".join(job.get("requirements", []) or [])),
        ("Benefits", ", ".join(job.get("benefits", []) or [])),
        ("Status", "Active" if job.get("is_active") else "Inactive"),
        ("Created At", str(job.get("created_at", "") or "")),
        ("Updated At", str(job.get("updated_at", "") or "")),
    ]

    ws.cell(row=1, column=1, value="Field")
    ws.cell(row=1, column=2, value="Value")
    _style_header_row(ws, 2)

    for i, (field, value) in enumerate(fields, start=2):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=2, value=value)

    _style_body_cells(ws, len(fields) + 1, 2)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80


def build_applicants_sheet(wb: Workbook, applicants: list[dict]) -> None:
    ws = wb.create_sheet(title="Applicants")

    headers = ["ID", "Name", "Email", "Phone", "Cover Letter", "Resume URL", "Status", "Applied At"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, len(headers))

    for i, app in enumerate(applicants, start=2):
        ws.cell(row=i, column=1, value=app.get("id", ""))
        ws.cell(row=i, column=2, value=app.get("name", ""))
        ws.cell(row=i, column=3, value=app.get("email", ""))
        ws.cell(row=i, column=4, value=app.get("phone", ""))
        ws.cell(row=i, column=5, value=app.get("cover_letter", ""))
        ws.cell(row=i, column=6, value=app.get("resume_url", ""))
        ws.cell(row=i, column=7, value=app.get("status", ""))
        ws.cell(row=i, column=8, value=str(app.get("created_at", "")))

    _style_body_cells(ws, len(applicants) + 1, len(headers))
    _auto_width(ws, len(headers))
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["E"].width = 50


def fetch_applications() -> list[dict]:
    logger.info("Fetching all applications from Supabase for export")
    try:
        with httpx.Client(timeout=_TIMEOUT) as client:
            response = client.get(_APPLICATIONS_ENDPOINT)
            response.raise_for_status()
            data = response.json()
            if isinstance(data, dict):
                return data.get("data", data.get("applications", []))
            if isinstance(data, list):
                return data
            return []
    except Exception as exc:
        logger.error("Failed to fetch applications for export: %s", str(exc))
        return []


def generate_excel_bytes(job: dict, applicants: list[dict]) -> io.BytesIO:
    wb = Workbook()
    build_job_sheet(wb, job)
    build_applicants_sheet(wb, applicants)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
